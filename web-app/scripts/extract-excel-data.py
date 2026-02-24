# Extract EUC Perth Assets from Excel and generate TypeScript seed data.

import json
import re
from datetime import datetime
from pathlib import Path

import openpyxl

EXCEL_PATH = r"C:\Users\Hard-Worker\Downloads\EUC_Perth_Assets.xlsx"
OUTPUT_PATH = Path(__file__).parent.parent / "src" / "data" / "seed.ts"

LOCATION_MAP = {
    '4.2': 'basement-4.2',
    'BR': 'build-room',
    'Darwin': 'darwin',
    'L17': 'level-17',
    'B4.3': 'basement-4.3',
}

SHEET_LOCATION_MAP = {
    '4.2': 'basement-4.2',
    'BR': 'build-room',
    'Darwin': 'darwin',
    'L17': 'level-17',
    'B4.3': 'basement-4.3',
}


def clean_str(val):
    if val is None:
        return ''
    s = str(val)
    s = s.replace('\xa0', ' ')
    s = s.replace('\ufffd', '"')  # replacement char → inch mark
    s = s.replace('\u2033', '"')  # double prime → inch mark
    s = s.replace('\u201d', '"')  # right double quote → inch mark
    # collapse multiple spaces and strip
    s = re.sub(r'\s+', ' ', s).strip()
    return s


def to_int(val, default=0):
    if val is None:
        return default
    try:
        return int(val)
    except (ValueError, TypeError):
        return default


def make_id(prefix, idx):
    return f"{prefix}-{idx:04d}"


def parse_timestamp(val):
    if val is None:
        return ''
    if isinstance(val, datetime):
        return val.isoformat()
    s = str(val).strip()
    # Try common formats
    for fmt in ('%Y-%m-%d %H:%M:%S', '%d/%m/%Y %H:%M:%S', '%d/%m/%Y %H:%M',
                '%Y-%m-%d %H:%M', '%d/%m/%Y', '%Y-%m-%d'):
        try:
            return datetime.strptime(s, fmt).isoformat()
        except ValueError:
            continue
    return s


def parse_action(action_str):
    """Parse action strings like 'add 3', 'subtract 1', 'add' into (action, volume)."""
    s = clean_str(action_str).lower()
    if not s:
        return 'add', 1

    match = re.match(r'(add|subtract)\s*(\d+)?', s)
    if match:
        action = match.group(1)
        volume = int(match.group(2)) if match.group(2) else 1
        return action, volume

    if 'subtract' in s or 'sub' in s:
        nums = re.findall(r'\d+', s)
        return 'subtract', int(nums[0]) if nums else 1
    if 'add' in s:
        nums = re.findall(r'\d+', s)
        return 'add', int(nums[0]) if nums else 1

    return 'add', 1


def map_location(loc_str):
    if loc_str is None:
        return 'basement-4.2'  # default
    s = clean_str(loc_str)
    return LOCATION_MAP.get(s, 'basement-4.2')


def extract_data():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    assets = []
    transactions = []
    san_records = []
    san_returns = []

    asset_id = 0
    txn_id = 0
    ret_id = 0

    # --- Extract Items from each location sheet ---
    for sheet_prefix, location_id in SHEET_LOCATION_MAP.items():
        sheet_name = f"{sheet_prefix}_Items"
        if sheet_name not in wb.sheetnames:
            print(f"  Skipping missing sheet: {sheet_name}")
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(min_row=1, values_only=True))
        if not rows:
            continue

        header = [clean_str(h).lower() for h in rows[0]]
        has_threshold = 'threshold' in header

        for row in rows[1:]:
            if not row or not row[0]:
                continue
            item_name = clean_str(row[0])
            if not item_name:
                continue

            raw_threshold = to_int(row[3] if has_threshold and len(row) > 3 else None, default=0)
            new_count = to_int(row[2] if len(row) > 2 else None)

            # Smart threshold: if Excel has 0, derive from stock level
            if raw_threshold == 0 and new_count > 0:
                threshold = max(3, round(new_count * 0.2))
            elif raw_threshold == 0:
                threshold = 5  # minimal default for zero-stock items
            else:
                threshold = raw_threshold

            asset_id += 1
            assets.append({
                'id': make_id('asset', asset_id),
                'item': item_name,
                'lastCount': to_int(row[1] if len(row) > 1 else None),
                'newCount': new_count,
                'threshold': threshold,
                'location': location_id,
            })

    # --- Extract Transactions from Timestamps sheets ---
    for sheet_prefix, location_id in SHEET_LOCATION_MAP.items():
        sheet_name = f"{sheet_prefix}_Timestamps"
        if sheet_name not in wb.sheetnames:
            print(f"  Skipping missing sheet: {sheet_name}")
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(min_row=1, values_only=True))
        if len(rows) < 2:
            continue

        for row in rows[1:]:
            if not row or not row[0]:
                continue
            timestamp = parse_timestamp(row[0])
            item_name = clean_str(row[1]) if len(row) > 1 else ''
            action_str = clean_str(row[2]) if len(row) > 2 else ''
            san_number = clean_str(row[3]) if len(row) > 3 else ''

            if not item_name:
                continue

            action, volume = parse_action(action_str)

            txn_id += 1
            txn = {
                'id': make_id('txn', txn_id),
                'timestamp': timestamp,
                'item': item_name,
                'action': action,
                'volume': volume,
                'location': location_id,
            }
            if san_number:
                txn['sanNumber'] = san_number
            transactions.append(txn)

    # --- Extract All SANs ---
    if 'All_SANs' in wb.sheetnames:
        ws = wb['All_SANs']
        rows = list(ws.iter_rows(min_row=1, values_only=True))
        for row in rows[1:]:
            if not row or not row[0]:
                continue
            san_number = clean_str(row[0])
            item = clean_str(row[1]) if len(row) > 1 else ''
            timestamp = parse_timestamp(row[2] if len(row) > 2 else None)
            location = map_location(row[3] if len(row) > 3 else None)

            if not san_number:
                continue

            san_records.append({
                'sanNumber': san_number,
                'item': item,
                'timestamp': timestamp,
                'location': location,
            })

    # --- Extract SAN Returns ---
    if 'SAN_Returns' in wb.sheetnames:
        ws = wb['SAN_Returns']
        rows = list(ws.iter_rows(min_row=1, values_only=True))
        for row in rows[1:]:
            if not row or not row[0]:
                continue
            san = clean_str(row[0])
            gen = clean_str(row[1]) if len(row) > 1 else 'G10'
            returned_by = clean_str(row[2]) if len(row) > 2 else ''
            returned_to = clean_str(row[3]) if len(row) > 3 else ''
            # row[4] is blank column
            notes = clean_str(row[5]) if len(row) > 5 else ''
            timestamp = parse_timestamp(row[6] if len(row) > 6 else None)

            if not san:
                continue

            # Validate generation value
            valid_gens = ['G5', 'G6', 'G7', 'G8', 'G9', 'G10', 'G11']
            if gen not in valid_gens:
                gen = 'G10'

            ret_id += 1
            san_returns.append({
                'id': make_id('ret', ret_id),
                'sanNumber': san,
                'generation': gen,
                'returnedBy': returned_by,
                'returnedTo': returned_to,
                'notes': notes,
                'timestamp': timestamp,
            })

    # Sort transactions by timestamp (newest first)
    transactions.sort(key=lambda t: t.get('timestamp', ''), reverse=True)

    return assets, san_records, san_returns, transactions


def generate_typescript(assets, san_records, san_returns, transactions):
    """Generate a TypeScript seed data module."""
    data = {
        'assets': assets,
        'sans': san_records,
        'returns': san_returns,
        'transactions': transactions,
    }

    # Pretty-print JSON for readability
    json_str = json.dumps(data, indent=2, ensure_ascii=False)

    ts_content = f"""import type {{ Asset, SANRecord, SANReturn, TransactionLog }} from '@/types';

interface SeedData {{
  assets: Asset[];
  sans: SANRecord[];
  returns: SANReturn[];
  transactions: TransactionLog[];
}}

export const seedData: SeedData = {json_str};
"""
    return ts_content


def main():
    print("Extracting data from Excel...")
    assets, san_records, san_returns, transactions = extract_data()

    print(f"  Assets: {len(assets)}")
    print(f"  SAN Records: {len(san_records)}")
    print(f"  SAN Returns: {len(san_returns)}")
    print(f"  Transactions: {len(transactions)}")

    # Show location breakdown
    from collections import Counter
    loc_counts = Counter(a['location'] for a in assets)
    for loc, count in sorted(loc_counts.items()):
        print(f"    {loc}: {count} items")

    print(f"\nGenerating TypeScript seed data...")
    ts_content = generate_typescript(assets, san_records, san_returns, transactions)

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(ts_content, encoding='utf-8')
    print(f"Written to: {OUTPUT_PATH}")
    print(f"File size: {OUTPUT_PATH.stat().st_size / 1024:.1f} KB")


if __name__ == '__main__':
    main()
