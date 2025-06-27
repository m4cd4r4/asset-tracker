# Macdara O Murchu
# 19.11.24


import logging.config
from pathlib import Path
from tkinter import Menu
import customtkinter as ctk
import os
import tkinter as tk
from tkinter import ttk
from openpyxl import load_workbook, Workbook
from datetime import datetime
import subprocess
from tkinter import filedialog
from tkinter import messagebox
import pandas as pd  # Ensure pandas is imported for date operations


# Function to save the workbook path to config.py
def save_config(workbook_path):
    with open('config.py', 'w') as config_file:
        config_file.write(f"workbook_path = r'{workbook_path}'\n")


logging_conf_path = Path('logging.conf')
if logging_conf_path.exists() and logging_conf_path.stat().st_size > 0:
    try:
        logging.config.fileConfig(logging_conf_path)
    except Exception as e:
        logging.error(f"Error configuring logging: {e}", exc_info=True)
else:
    logging.basicConfig(level=logging.DEBUG)


def run_inventory_script(script_name, output_prefix, success_message):
    """
    Generalized function to run an inventory script and handle its output.
    """
    script_path = script_directory / script_name
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")  # Generate a timestamp
    output_path = script_directory / "Plots" / f"{output_prefix}_{timestamp}.png"

    # Ensure the output directory exists
    output_path.parent.mkdir(parents=True, exist_ok=True)

    if script_path.exists():
        try:
            subprocess.run(
                ["python", str(script_path), "--output", str(output_path)],
                check=True,
            )
            messagebox.showinfo("Success", f"{success_message} saved to {output_path}")
        except subprocess.CalledProcessError as e:
            logging.error(f"Error running {script_name}: {e}")
            tk.messagebox.showerror("Error", f"Failed to run the script: {e}")
    else:
        tk.messagebox.showerror(
            "Error", f"The script '{script_name}' does not exist in the directory."
        )


def set_item_threshold():
    """
    Open a dialog to allow users to set custom thresholds for items.
    """
    threshold_window = tk.Toplevel(root)
    threshold_window.title("Set Item Threshold")
    threshold_window.geometry("500x400")

    # Select sheet dropdown
    tk.Label(threshold_window, text="Select Sheet:").pack(pady=5)
    sheet_var = tk.StringVar(value='4.2_Items')
    sheet_dropdown = ttk.Combobox(threshold_window, textvariable=sheet_var, values=['4.2_Items', 'BR_Items', 'Darwin_Items'])
    sheet_dropdown.pack(pady=5)

    # Select item dropdown
    tk.Label(threshold_window, text="Select Item:").pack(pady=5)
    item_var = tk.StringVar()
    item_dropdown = ttk.Combobox(threshold_window, textvariable=item_var)
    item_dropdown.pack(pady=5)

    # Update item list when sheet changes
    def update_item_list(*args):
        sheet_name = sheet_var.get()
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            items = [row[0] for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True) if row[0]]
            item_dropdown['values'] = items

    sheet_var.trace("w", update_item_list)
    update_item_list()

    # Threshold input
    tk.Label(threshold_window, text="Set Threshold:").pack(pady=5)
    threshold_var = tk.IntVar(value=10)
    threshold_entry = ttk.Entry(threshold_window, textvariable=threshold_var)
    threshold_entry.pack(pady=5)

    # Save threshold button
    def save_threshold():
        sheet_name = sheet_var.get()
        item_name = item_var.get()
        threshold_value = threshold_var.get()

        if not sheet_name or not item_name or threshold_value is None:
            tk.messagebox.showerror("Error", "All fields are required.", parent=threshold_window)
            return

        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0] == item_name:
                    row_idx = row[0].row
                    threshold_col_idx = sheet.max_column  # Assume last column is Threshold
                    sheet.cell(row=row_idx, column=threshold_col_idx, value=threshold_value)
                    workbook.save(workbook_path)
                    tk.messagebox.showinfo("Success", f"Threshold for '{item_name}' set to {threshold_value}.", parent=threshold_window)
                    break
            else:
                tk.messagebox.showerror("Error", f"Item '{item_name}' not found in {sheet_name}.", parent=threshold_window)

    tk.Button(threshold_window, text="Save Threshold", command=save_threshold).pack(pady=10)


# Use the generalized function for specific scripts
def run_basement_4_2_inventory():
    run_inventory_script(
        script_name="inventory-levels_4.2v3.py",
        output_prefix="Basement_4.2_inventory",
        success_message="Basement 4.2 inventory plot"
    )

def run_build_room_inventory():
    run_inventory_script(
        script_name="inventory-levels_BRv2.3.py",
        output_prefix="Build_Room_inventory",
        success_message="Build Room inventory plot"
    )

def run_darwin_inventory():
    run_inventory_script(
        script_name="inventory-levels_darwin.v2.py",
        output_prefix="Darwin_inventory",
        success_message="Darwin inventory plot"
    )

def run_combined_inventory():
    run_inventory_script(
        script_name="inventory-levels_combinedv1.5.py",
        output_prefix="Combined_inventory",
        success_message="Combined inventory plot"
    )


def save_plots():
    """
    Saves an image of each inventory data to a timestamped subfolder within the 'Plots' directory.
    """
    # Create a timestamped subfolder within the 'Plots' directory
    plots_dir = script_directory / "Plots"
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    save_dir = plots_dir / f"All_{timestamp}"
    save_dir.mkdir(parents=True, exist_ok=True)  # Ensure the directory exists

    # Define the scripts and their corresponding output prefixes and success messages
    inventory_scripts = [
        ("inventory-levels_4.2v3.py", "Basement_4.2_inventory", "Basement 4.2 inventory plot"),
        ("inventory-levels_BRv2.3.py", "Build_Room_inventory", "Build Room inventory plot"),
        ("inventory-levels_darwin.v2.py", "Darwin_inventory", "Darwin inventory plot"),
        ("inventory-levels_combinedv1.5.py", "Combined_inventory", "Combined inventory plot"),
    ]

    # Loop through each script and generate plots
    for script_name, output_prefix, success_message in inventory_scripts:
        script_path = script_directory / script_name
        if script_path.exists():
            try:
                output_path = save_dir / f"{output_prefix}.png"
                subprocess.run(
                    ["python", str(script_path), "--output", str(output_path)],
                    check=True,
                    env={**os.environ, "MPLBACKEND": "Agg"}  # Headless matplotlib
                )
                logging.info(f"{success_message} saved to {output_path}")
            except subprocess.CalledProcessError as e:
                logging.error(f"Error while saving {success_message}: {e}")
                tk.messagebox.showerror("Error", f"Error while saving {success_message}: {e}")
        else:
            logging.error(f"Script '{script_name}' not found!")
            tk.messagebox.showerror("Error", f"Script '{script_name}' not found!")

    # Open the folder with saved plots
    open_folder_prompt(save_dir)


def open_folder_prompt(folder_path):
    """
    Prompts the user to open the folder where plots are saved.
    """
    def open_folder():
        if os.name == 'nt':  # Windows
            os.startfile(folder_path)
        elif sys.platform == 'darwin':  # macOS
            subprocess.run(['open', folder_path])
        else:  # Linux and others
            subprocess.run(['xdg-open', folder_path])

    info_window = tk.Toplevel(root)
    info_window.title("Info")
    info_window.geometry("400x150")
    tk.Label(info_window, text=f"All plots have been saved in:\n{folder_path}", wraplength=350).pack(pady=10)

    open_folder_button = ttk.Button(info_window, text="Open folder", command=open_folder)
    open_folder_button.pack(pady=10)



def open_spreadsheet():
    try:
        if os.name == 'nt':
            os.startfile(workbook_path)
        else:
            opener = "open" if sys.platform == "darwin" else "xdg-open"
            subprocess.run([opener, workbook_path])
    except Exception as e:
        tk.messagebox.showerror("Error", f"Failed to open the spreadsheet: {e}")


def update_all_sans_location():
    """
    Updates the 'Location' column in the 'All_SANs' sheet based on the presence of SANs
    in specific timestamp sheets.
    """
    # Ensure the 'All_SANs' sheet exists
    if 'All_SANs' not in workbook.sheetnames:
        tk.messagebox.showerror("Error", "'All_SANs' sheet not found in the workbook.")
        return

    all_sans_sheet = workbook['All_SANs']

    # Ensure the Location column (D) has a header
    if all_sans_sheet.max_column < 4:
        all_sans_sheet.cell(row=1, column=4, value="Location")

    # Mapping of sheets to short locations
    location_mapping = {
        '4.2_Timestamps': '4.2',
        'BR_Timestamps': 'BR',
        'Darwin_Timestamps': 'Darwin'
    }

    # Iterate over rows in All_SANs and update the location
    for row_idx, row in enumerate(all_sans_sheet.iter_rows(min_row=2, values_only=True), start=2):
        san_number = row[0]
        location = None

        # Check each timestamp sheet for the SAN
        for sheet_name, loc in location_mapping.items():
            if sheet_name in workbook.sheetnames:
                timestamp_sheet = workbook[sheet_name]
                if any(san_number == cell[0] for cell in timestamp_sheet.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True)):
                    location = loc
                    break

        # Write the location to the Location column (Column D)
        all_sans_sheet.cell(row=row_idx, column=4, value=location)

    # Save the workbook after updating
    workbook.save(workbook_path)


def view_all_sans_log():
    """
    Displays the updated All_SANs data in a Treeview widget with columns
    'SAN Number', 'Item', 'Time', and 'Location'.
    Includes a search box to filter SAN numbers dynamically.
    """
    # Update the 'Location' column in the 'All_SANs' sheet
    update_all_sans_location()

    log_window = tk.Toplevel(root)
    log_window.title("SANs In Stock")
    log_window.geometry("800x800")

    # Search box frame
    search_frame = tk.Frame(log_window)
    search_frame.pack(fill="x", padx=10, pady=5)

    tk.Label(search_frame, text="Search:").pack(side="left", padx=5)
    search_var = tk.StringVar()

    search_entry = ttk.Entry(search_frame, textvariable=search_var)
    search_entry.pack(side="left", fill="x", expand=True, padx=5)

    # Treeview and scrollbar
    columns = ("SAN Number", "Item", "Time", "Location")
    log_tree = ttk.Treeview(log_window, columns=columns, show="headings")
    for col in columns:
        log_tree.heading(col, text=col)
        log_tree.column(col, anchor="w")
    log_tree.pack(expand=True, fill="both", padx=10, pady=10)

    scrollbar = ttk.Scrollbar(log_window, orient="vertical", command=log_tree.yview)
    scrollbar.pack(side="right", fill="y")
    log_tree.configure(yscrollcommand=scrollbar.set)

    # Load data from the "All_SANs" sheet
    def load_data(filter_text=""):
        """
        Populates the Treeview with data, filtering by filter_text.
        """
        log_tree.delete(*log_tree.get_children())  # Clear current data
        if 'All_SANs' in workbook.sheetnames:
            all_sans_sheet = workbook['All_SANs']
            for row in all_sans_sheet.iter_rows(min_row=2, values_only=True):
                san_number, item, timestamp, location = row
                # Filter by the search text
                if filter_text.lower() in str(san_number).lower():
                    log_tree.insert('', 'end', values=(san_number, item, timestamp, location))
        else:
            tk.messagebox.showinfo("Info", "'All_SANs' sheet not found or empty.", parent=log_window)

    # Trigger search on text change
    def on_search(*args):
        filter_text = search_var.get()
        load_data(filter_text)

    search_var.trace("w", on_search)  # Bind dynamic updates to search

    # Load initial data
    load_data()

def view_san_returns_log():
    """
    Displays the 'SAN_Returns' sheet in a Treeview widget with columns:
    'SAN', 'Gen', 'Returned By', 'Returned To', 'Notes', and 'Timestamp'.
    """
    # Ensure the 'SAN_Returns' sheet exists
    if 'SAN_Returns' not in workbook.sheetnames:
        tk.messagebox.showinfo("Info", "'SAN_Returns' sheet not found. Creating it now.")
        san_returns_sheet = workbook.create_sheet('SAN_Returns')
        san_returns_sheet.append(["SAN", "Gen", "Returned By", "Returned To", "Notes", "Timestamp"])  # Updated headers
        workbook.save(workbook_path)

    log_window = tk.Toplevel(root)
    log_window.title("SAN Return Log")
    log_window.geometry("900x600")

    # Treeview for the SAN_Returns sheet
    columns = ("SAN", "Gen", "Returned By", "Returned To", "Notes", "Timestamp")  # Include Gen column
    returns_tree = ttk.Treeview(log_window, columns=columns, show="headings")
    for col in columns:
        returns_tree.heading(col, text=col)
        returns_tree.column(col, anchor="w")
    returns_tree.pack(expand=True, fill="both", padx=10, pady=10)

    scrollbar = ttk.Scrollbar(log_window, orient="vertical", command=returns_tree.yview)
    scrollbar.pack(side="right", fill="y")
    returns_tree.configure(yscrollcommand=scrollbar.set)

    # Load data from the "SAN_Returns" sheet
    refresh_san_returns_log(returns_tree)

    # Add the right-click copy option to the Treeview
    add_copy_option(returns_tree)

    # Add the "Return Asset" button at the bottom
    return_asset_button = ctk.CTkButton(
        log_window,
        text="Return Asset",
        font=("Helvetica", 14),
        command=lambda: open_san_return_form_with_tree(returns_tree)  # Pass returns_tree to the form
    )
    return_asset_button.pack(pady=10)  # Positioned at the bottom of the log window


def open_san_return_form_with_tree(returns_tree):
    """
    Opens the SAN Return Form and passes the Treeview reference for dynamic updates.
    """
    form_window = ctk.CTkToplevel(root)
    form_window.title("SAN Return Form")
    form_window.geometry("500x550")

    # Set a uniform appearance for the form
    ctk.set_appearance_mode("dark")  # Options: "dark", "light", "system"
    ctk.set_default_color_theme("blue")  # Options: "blue", "green", "dark-blue"

    # Title Label
    title_label = ctk.CTkLabel(
        form_window,
        text="SAN Return Form",
        font=("Helvetica", 20, "bold"),
    )
    title_label.pack(pady=20)

    # Frame for form fields
    form_frame = ctk.CTkFrame(
        form_window, corner_radius=10, fg_color="#2C2F33"
    )  # Darker background
    form_frame.pack(padx=20, pady=10, fill="both", expand=True)

    # Fields and Input
    fields = ["SAN", "Gen", "Returned By", "Returned To", "Notes", "Timestamp"]
    entries = {}
    gen_values = ["G5", "G6", "G7", "G8", "G9", "G10", "G11"]

    for idx, field in enumerate(fields):
        label = ctk.CTkLabel(
            form_frame, text=field, font=("Helvetica", 14), anchor="w"
        )
        label.grid(row=idx, column=0, padx=10, pady=10, sticky="e")

        if field == "Gen":
            # Dropdown for "Gen"
            entries[field] = ctk.CTkComboBox(
                form_frame, values=gen_values, width=250, corner_radius=10
            )
            entries[field].set(gen_values[0])  # Default to the first value
        elif field == "Timestamp":
            # Prefill Timestamp
            entries[field] = ctk.CTkEntry(
                form_frame,
                width=250,
                corner_radius=10,
                placeholder_text="YYYY-MM-DD HH:MM:SS",
            )
            entries[field].insert(0, datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        else:
            entries[field] = ctk.CTkEntry(
                form_frame, width=250, corner_radius=10, placeholder_text=f"Enter {field}"
            )

        entries[field].grid(row=idx, column=1, padx=10, pady=10)

    # Buttons Frame
    buttons_frame = ctk.CTkFrame(form_window, fg_color="#23272A", corner_radius=10)
    buttons_frame.pack(pady=20)

    # Submit Button
    submit_button = ctk.CTkButton(
        buttons_frame,
        text="Submit",
        font=("Helvetica", 14, "bold"),
        width=120,
        corner_radius=10,
        command=lambda: submit_san_return(
            entries["SAN"].get(),
            entries["Gen"].get(),
            entries["Returned By"].get(),
            entries["Returned To"].get(),
            entries["Notes"].get(),
            entries["Timestamp"].get(),
            form_window,
            returns_tree,  # Pass Treeview reference
        ),
    )
    submit_button.pack(side="left", padx=10)

    # Cancel Button
    cancel_button = ctk.CTkButton(
        buttons_frame,
        text="Cancel",
        font=("Helvetica", 14, "bold"),
        width=120,
        corner_radius=10,
        fg_color="red",
        command=form_window.destroy,
    )
    cancel_button.pack(side="left", padx=10)



def submit_san_return(san, gen, returned_by, returned_to, notes, timestamp, form_window, returns_tree):
    """
    Validates input and submits the data to the SAN_Returns sheet.
    Updates the Treeview dynamically after submission.
    """
    # Check if required fields are filled
    if not san or not gen or not returned_by or not returned_to or not timestamp:
        tk.messagebox.showerror("Error", "All fields except 'Notes' are required.", parent=form_window)
        return

    # Ensure the 'SAN_Returns' sheet exists
    if 'SAN_Returns' not in workbook.sheetnames:
        san_returns_sheet = workbook.create_sheet('SAN_Returns')
        san_returns_sheet.append(["SAN", "Gen", "Returned By", "Returned To", "Notes", "Timestamp"])  # Add headers
    else:
        san_returns_sheet = workbook['SAN_Returns']

    # Append the data to the SAN_Returns sheet
    san_returns_sheet.append([san, gen, returned_by, returned_to, notes, timestamp])

    # Save the workbook
    workbook.save(workbook_path)

    # Update the Treeview dynamically
    refresh_san_returns_log(returns_tree)

    # Inform the user and close the form
    tk.messagebox.showinfo("Success", "SAN return data submitted successfully!", parent=form_window)
    form_window.destroy()  # Close the form window


def refresh_san_returns_log(returns_tree):
    """
    Refreshes the Treeview with the latest data from the SAN_Returns sheet.
    """
    returns_tree.delete(*returns_tree.get_children())  # Clear current data
    if 'SAN_Returns' in workbook.sheetnames:
        san_returns_sheet = workbook['SAN_Returns']
        for row in san_returns_sheet.iter_rows(min_row=2, values_only=True):
            returns_tree.insert('', 'end', values=row)


def view_san_returns_log():
    """
    Displays the 'SAN_Returns' sheet in a Treeview widget with columns:
    'SAN', 'Returned By', 'Returned To', 'Notes', and 'Timestamp'.
    """
    # Ensure the 'SAN_Returns' sheet exists
    if 'SAN_Returns' not in workbook.sheetnames:
        tk.messagebox.showinfo("Info", "'SAN_Returns' sheet not found. Creating it now.")
        san_returns_sheet = workbook.create_sheet('SAN_Returns')
        san_returns_sheet.append(["SAN", "Returned By", "Returned To", "Notes", "Timestamp"])
        workbook.save(workbook_path)

    log_window = tk.Toplevel(root)
    log_window.title("SAN Return List")
    log_window.geometry("800x600")

    # Treeview for the SAN_Returns sheet
    columns = ("SAN", "Returned By", "Returned To", "Notes", "Timestamp")
    returns_tree = ttk.Treeview(log_window, columns=columns, show="headings")
    for col in columns:
        returns_tree.heading(col, text=col)
        returns_tree.column(col, anchor="w")
    returns_tree.pack(expand=True, fill="both", padx=10, pady=10)

    scrollbar = ttk.Scrollbar(log_window, orient="vertical", command=returns_tree.yview)
    scrollbar.pack(side="right", fill="y")
    returns_tree.configure(yscrollcommand=scrollbar.set)

    # Load data from the "SAN_Returns" sheet
    refresh_san_returns_log(returns_tree)

    # Add the right-click copy option to the Treeview
    add_copy_option(returns_tree)

    # Add the "Return Asset" button at the bottom
    return_asset_button = ctk.CTkButton(
        log_window,
        text="Return Asset",
        font=("Helvetica", 14),
        command=lambda: open_san_return_form_with_tree(returns_tree)  # Pass returns_tree to the form
    )
    return_asset_button.pack(pady=10)  # Positioned at the bottom of the log window


def open_san_return_form_with_tree(returns_tree):
    """
    Opens the SAN Return Form and passes the Treeview reference for dynamic updates.
    """
    form_window = tk.Toplevel(root)
    form_window.title("SAN Return Form")
    form_window.geometry("400x400")

    # Labels and Entry fields
    fields = ["SAN", "Returned By", "Returned To", "Notes", "Timestamp"]
    entries = {}

    for idx, field in enumerate(fields):
        tk.Label(form_window, text=field).grid(row=idx, column=0, padx=10, pady=5, sticky="e")
        entry = tk.Entry(form_window, width=30)
        entry.grid(row=idx, column=1, padx=10, pady=5)
        entries[field] = entry

    # Timestamp prefilled with current datetime
    entries["Timestamp"].insert(0, datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

    # Submit Button
    tk.Button(
        form_window,
        text="Submit",
        command=lambda: submit_san_return(
            entries["SAN"].get(),
            entries["Returned By"].get(),
            entries["Returned To"].get(),
            entries["Notes"].get(),
            entries["Timestamp"].get(),
            form_window,
            returns_tree  # Pass the Treeview to the submit function
        )
    ).grid(row=len(fields), column=0, columnspan=2, pady=20)


root = ctk.CTk()
root.title("EUC Assets - WA")
root.geometry("675x850")

# Unified font settings for the application
custom_font = ("Helvetica", 12)

# Custom styles for Treeview
style = ttk.Style()
style.theme_use("clam")
style.configure("Treeview",
                font=("Helvetica", 12),
                rowheight=25,
                background="#f9f9f9",
                foreground="#333",
                fieldbackground="#f9f9f9")
style.map("Treeview", background=[("selected", "#3399ff")])



menu_bar = tk.Menu(root)

# Create a submenu for SANs
sans_menu = tk.Menu(menu_bar, tearoff=0)
sans_menu.add_command(label="SANs In Stock", command=view_all_sans_log)
sans_menu.add_command(label="Returned SANs - List", command=view_san_returns_log)
sans_menu.add_command(label="Returned SANs", command=lambda: open_san_return_form_with_tree(None))

# Create a submenu for Inventory
inventory_menu = tk.Menu(menu_bar, tearoff=0)
inventory_menu.add_command(label="Basement 4.2 Inventory", command=run_basement_4_2_inventory)
inventory_menu.add_command(label="Build Room Inventory", command=run_build_room_inventory)  # Corrected function name
inventory_menu.add_command(label="Darwin Inventory", command=run_darwin_inventory)
inventory_menu.add_command(label="Combined Inventory", command=run_combined_inventory)
inventory_menu.add_command(label="Save Plots", command=save_plots)


# Create the main Options menu
plots_menu = tk.Menu(menu_bar, tearoff=0)
plots_menu.add_cascade(label="SANs", menu=sans_menu)  # Add SANs submenu
plots_menu.add_cascade(label="Inventory", menu=inventory_menu)  # Add Inventory submenu
plots_menu.add_command(label="Open Spreadsheet", command=open_spreadsheet)
plots_menu.add_command(label="Check Restock Threshold", command=lambda: check_restock_threshold(10))
# plots_menu.add_command(label="Headsets In Stock", command=view_headsets_log)

menu_bar.add_cascade(label="Options", menu=plots_menu)
root.config(menu=menu_bar)



script_directory = Path(__file__).parent

def get_file_path():
    # Creating a temporary root window for file dialog
    temp_root = tk.Tk()
    temp_root.withdraw()  # Hide the temporary root window
    file_path = filedialog.askopenfilename(
        title="Select a spreadsheet file",
        filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*"))
    )
    temp_root.destroy()  # Destroy the temporary root window to clean up
    if not file_path:
        tk.messagebox.showerror("Error", "No file selected. Exiting application.")
        raise SystemExit  # Exit the application if no file is selected
    return file_path

# Get the workbook path from user
workbook_path = get_file_path()
save_config(workbook_path)  # Save the path to the config file immediately after getting it
workbook = load_workbook(workbook_path)

def ensure_threshold_column():
    """
    Ensure each inventory sheet has a 'Threshold' column. Add it if missing.
    """
    try:
        for sheet_name in ['4.2_Items', 'BR_Items', 'Darwin_Items']:
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                max_col = sheet.max_column
                header_row = [sheet.cell(row=1, column=col).value for col in range(1, max_col + 1)]

                if 'Threshold' not in header_row:
                    sheet.cell(row=1, column=max_col + 1, value='Threshold')
                    for row in range(2, sheet.max_row + 1):
                        sheet.cell(row=row, column=max_col + 1, value=10)  # Default threshold value
                    workbook.save(workbook_path)
                    logging.info(f"Added 'Threshold' column to {sheet_name}.")
    except Exception as e:
        logging.error(f"Error ensuring 'Threshold' column: {e}")
        tk.messagebox.showerror("Error", f"Failed to add 'Threshold' column: {e}")


all_sans_sheet = workbook['All_SANs']
sheets = {
    'original': ('4.2_Items', '4.2_Timestamps'),
    'backup': ('BR_Items', 'BR_Timestamps'),
    'L17': ('L17_Items', 'L17_Timestamps'),
    'B4.3': ('B4.3_Items', 'B4.3_Timestamps'),
    'Darwin': ('Darwin_Items', 'Darwin_Timestamps')
}

current_sheets = sheets['original']

style = ttk.Style()
style.configure("Treeview", font=('Helvetica', 12,))

vcmd = (root.register(lambda P: P.isdigit() or P == ""), '%P')
class SANInputDialog(tk.Toplevel):
    def __init__(self, parent, title=None):
        super().__init__(parent)
        self.transient(parent)
        self.title(title)
        self.parent = parent
        self.result = None
        self.create_widgets()
        self.grab_set()
        # Center the dialog window on the parent
        self.geometry(f"+{parent.winfo_rootx() + parent.winfo_width() // 2 - 100}+{parent.winfo_rooty() + parent.winfo_height() // 2 - 50}")
        self.after(10, self.focus_entry)  # Ensure the entry field gets focus
        self.wait_window(self)

    def create_widgets(self):
        # Input field
        self.entry = ttk.Entry(self, validate="key", validatecommand=vcmd)
        self.entry.pack(padx=5, pady=5)
        self.entry.bind("<Return>", lambda _: self.on_submit())  # Bind Enter key to submit action

        # Buttons frame
        button_frame = tk.Frame(self)
        button_frame.pack(pady=5)

        # Submit button
        submit_button = ttk.Button(button_frame, text="Submit", command=self.on_submit)
        submit_button.pack(side='left', padx=5)

        # Cancel button
        cancel_button = ttk.Button(button_frame, text="Cancel", command=self.on_cancel)
        cancel_button.pack(side='left', padx=5)

    def focus_entry(self):
        """
        Force the focus to the entry field.
        """
        self.entry.focus_force()

    def on_submit(self):
        san_input = self.entry.get()
        if san_input and len(san_input) >= 5 and len(san_input) <= 6:
            self.result = san_input
            self.destroy()
        else:
            tk.messagebox.showerror("Error", "Please enter a valid SAN number.", parent=self)
            self.after(10, self.focus_entry)  # Reapply focus after showing an error

    def on_cancel(self):
        self.result = None
        self.destroy()




def is_san_unique(san_number):
    # Adjust the search to account for the 'SAN' prefix properly
    search_string = "SAN" + san_number if not san_number.startswith("SAN") else san_number
    unique = all(search_string != row[0] for row in all_sans_sheet.iter_rows(min_row=2, values_only=True))
    print(f"Checking SAN {search_string}: Unique - {unique}")  # Debug print
    return unique


def show_san_input():
    dialog = SANInputDialog(root, "Enter SAN Number")
    return dialog.result

frame = ctk.CTkFrame(root)
frame.pack(padx=3, pady=3, fill='both', expand=True)

button_width = 120  # Define the width of the buttons

entry_frame = ctk.CTkFrame(frame)
entry_frame.pack(pady=3, fill='x')  # Fill horizontally for spacing alignment

# Button frame for location buttons
location_buttons_frame = ctk.CTkFrame(entry_frame)
location_buttons_frame.pack(fill='x', padx=10, pady=5)  # Expand frame to fill horizontally

# Use grid to align and space buttons evenly
buttons = [
    ("Basement 4.2", lambda: switch_sheets('original')),
    ("Build Room", lambda: switch_sheets('backup')),
    ("Darwin", lambda: switch_sheets('Darwin')),
    ("Level 17", lambda: switch_sheets('L17')),
    ("Basement 4.3", lambda: switch_sheets('B4.3')),
]

# Variable to store the current selected button
selected_button = None

def highlight_button(selected, buttons):
    """
    Highlights the selected button by making its label bold and resets others.
    
    :param selected: The selected button widget.
    :param buttons: List of all button widgets.
    """
    global selected_button
    for btn in buttons:
        if btn == selected:
            btn.configure(font=("Helvetica", 14, "bold"))  # Set font to bold
        else:
            btn.configure(font=("Helvetica", 14, "normal"))  # Reset to normal
    selected_button = selected

# Create the buttons and store references in a list
button_widgets = []

# Variable to store the currently selected button
selected_button = None

def highlight_button(selected, buttons):
    """
    Highlights the selected button by making its label bold and resets others.
    
    :param selected: The selected button widget.
    :param buttons: List of all button widgets.
    """
    global selected_button
    for btn in buttons:
        if btn == selected:
            btn.configure(font=("Helvetica", 14, "bold"))  # Set font to bold
        else:
            btn.configure(font=("Helvetica", 14, "normal"))  # Reset to normal
    selected_button = selected

# Create the buttons and store references in a list
button_widgets = []

for col, (text, command) in enumerate(buttons):
    # Define the button outside the lambda
    btn = ctk.CTkButton(
        location_buttons_frame,
        text=text,
        width=button_width,
        font=("Helvetica", 14, "normal"),  # Default font is normal
        corner_radius=3
    )
    btn.configure(
        command=lambda b=btn, cmd=command: [highlight_button(b, button_widgets), root.after(1, cmd)]
    )
    btn.grid(row=0, column=col, padx=5, pady=5, sticky="ew")
    button_widgets.append(btn)


def update_treeview():
    tree.delete(*tree.get_children())
    workbook = load_workbook(workbook_path)
    item_sheet = workbook[current_sheets[0]]
    row_count = 0
    for row in item_sheet.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            if row_count % 2 == 0:
                bg_color = 'white'
            else:
                bg_color = '#f0f0f0'
            tree.insert('', 'end', values=row, tags=('oddrow' if row_count % 2 == 1 else 'evenrow'))
            tree.tag_configure('oddrow', background='#f0f0f0')
            tree.tag_configure('evenrow', background='white')
            row_count += 1

def log_change(item, action, san_number="", timestamp_sheet=None, volume=1):  # Added volume parameter with default value of 1
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    try:
        if timestamp_sheet is not None:
            san_number = f"SAN{san_number}" if san_number and not san_number.startswith('SAN') else san_number
            
            # Conditionally modify the action string to include volume for non-SAN items
            if san_number == "":
                action_text = f"{action} {volume}"
            else:
                action_text = action
            
            timestamp_sheet.append([timestamp, item, action_text, san_number])  # Use action_text instead of action
            workbook.save(workbook_path)
            update_log_view()
            logging.info(f"Logged change: Time: {timestamp}, Item: {item}, Action: {action_text}, SAN: {san_number}")  # Use action_text
        else:
            logging.error("No timestamp sheet provided for logging.")
    except Exception as e:
        logging.error(f"Failed to log change: {e}")
        tk.messagebox.showerror("Error", f"Failed to log change: {e}")

def switch_sheets(sheet_type):
    global current_sheets
    current_sheets = sheets[sheet_type]
    update_treeview()
    update_log_view()

def update_log_view():
    if 'log_view' in globals():
        log_view.delete(*log_view.get_children())
        log_sheet = workbook[current_sheets[1]]
        all_rows = list(log_sheet.iter_rows(min_row=2, values_only=True))
        # Adjust the sorting to use the first column (timestamp)
        sorted_rows = sorted(all_rows, key=lambda r: datetime.strptime(r[0], "%Y-%m-%d %H:%M:%S") if r[0] else datetime.min, reverse=True)
        row_count = 0
        for row in sorted_rows:
            if row[0] is not None:
                log_view.insert('', 'end', values=row, tags=('oddrow' if row_count % 2 == 1 else 'evenrow'))
                log_view.tag_configure('oddrow', background='#f0f0f0')
                log_view.tag_configure('evenrow', background='white')
                row_count += 1


def update_count(operation):
    """
    Updates the inventory count and handles SAN addition/removal, including logging
    and updating the location in the 'All_SANs' sheet.
    """
    selected_item = tree.item(tree.focus())['values'][0] if tree.focus() else None
    if selected_item:
        input_value = entry_value.get()
        if input_value.isdigit():
            input_value = int(input_value)
            item_sheet = workbook[current_sheets[0]]
            timestamp_sheet = workbook[current_sheets[1]]
            san_required = any(g in selected_item for g in ["G8", "G9", "G10"])

            entered_sans_count = 0  # To keep track of successfully entered SAN numbers if needed

            if san_required:
                # Loop to input each SAN for SAN-required items
                while entered_sans_count < input_value:
                    # Small delay to ensure dialog focus is handled correctly
                    root.after(50)  # Ensures the UI thread processes events between dialogs
                    san_number = show_san_input()
                    if san_number is None:  # User cancelled the input
                        break  # Keep the already entered SANs

                    # Ensure SAN number has the 'SAN' prefix
                    san_number = "SAN" + san_number if not san_number.startswith("SAN") else san_number
                    
                    # Determine the location of the SAN based on the current sheet
                    location_mapping = {
                        '4.2_Timestamps': '4.2',
                        'BR_Timestamps': 'BR',
                        'Darwin_Timestamps': 'Darwin'
                    }
                    current_location = None
                    for sheet_name, loc in location_mapping.items():
                        if current_sheets[1] == sheet_name:
                            current_location = loc
                            break

                    if operation == 'add':
                        if is_san_unique(san_number):
                            # Append the SAN, Item, Timestamp, and Location to the "All_SANs" sheet
                            all_sans_sheet.append([san_number, selected_item, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), current_location])
                            # Log each SAN unique number immediately
                            log_change(selected_item, operation, san_number, timestamp_sheet, volume=1)
                            workbook.save(workbook_path)
                            entered_sans_count += 1
                        else:
                            tk.messagebox.showerror("Error", "Duplicate or already used SAN number.", parent=root)
                    elif operation == 'subtract':
                        # Search for the SAN in all_sans_sheet and remove if found and matches the item
                        for row in all_sans_sheet.iter_rows(min_row=2):
                            if row[0].value == san_number and row[1].value == selected_item:
                                all_sans_sheet.delete_rows(row[0].row)
                                log_change(selected_item, operation, san_number, timestamp_sheet, volume=1)
                                workbook.save(workbook_path)
                                entered_sans_count += 1
                                break
                        else:  # Executed if the loop completes without breaking (SAN not found or doesn't match item)
                            tk.messagebox.showerror("Error", f"SAN number {san_number} does not match the selected item.", parent=root)

            for row in item_sheet.iter_rows(min_row=2):
                if row[0].value == selected_item:
                    row[1].value = row[2].value or 0  # Update LastCount to the current NewCount
                    if operation == 'add':
                        row[2].value = (row[2].value or 0) + (input_value if not san_required else entered_sans_count)
                    elif operation == 'subtract':
                        row[2].value = max((row[2].value or 0) - (input_value if not san_required else entered_sans_count), 0)

            # Log volume for non-SAN items or after all entered SANs
            if (san_required and entered_sans_count > 0) or not san_required:
                volume_to_log = input_value if not san_required else entered_sans_count
                log_change(selected_item, operation, "", timestamp_sheet, volume=volume_to_log)

            workbook.save(workbook_path)
            update_treeview()
            update_log_view()

        # **Add this line to refocus on the entry field after processing**
        entry_value.focus_set()

def check_restock_threshold(threshold=10):
    """
    Display items that need restocking based on individual thresholds.
    """
    try:
        low_stock_items = []

        # Iterate through sheets to check against thresholds
        for sheet_name in ['4.2_Items', 'BR_Items', 'Darwin_Items']:
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    item, _, new_count, item_threshold = row[:4]  # Adjust columns as necessary
                    if new_count is not None and item_threshold is not None and new_count < item_threshold:
                        low_stock_items.append((sheet_name, item, new_count, item_threshold))

        # Display results
        if low_stock_items:
            restock_window = tk.Toplevel(root)
            restock_window.title("Low Stock Alert")
            restock_window.geometry("500x400")

            tk.Label(restock_window, text="Items Below Threshold", font=("Helvetica", 14, "bold")).pack(pady=10)

            columns = ("Sheet", "Item", "Stock", "Threshold")
            restock_tree = ttk.Treeview(restock_window, columns=columns, show="headings", height=15)
            for col in columns:
                restock_tree.heading(col, text=col)
                restock_tree.column(col, anchor='w', width=150 if col == "Item" else 100)

            restock_tree.pack(expand=True, fill="both", padx=10, pady=10)

            for sheet, item, count, item_threshold in low_stock_items:
                restock_tree.insert("", "end", values=(sheet, item, count, item_threshold))

            scrollbar = ttk.Scrollbar(restock_window, orient="vertical", command=restock_tree.yview)
            scrollbar.pack(side="right", fill="y")
            restock_tree.configure(yscrollcommand=scrollbar.set)
        else:
            tk.messagebox.showinfo("Info", "No items are below their thresholds.")
    except Exception as e:
        logging.error(f"Failed to check restock threshold: {e}")
        tk.messagebox.showerror("Error", f"Failed to check restock threshold: {e}")


# Treeview for items
columns = ("Item", "LastCount", "NewCount")
tree = ttk.Treeview(frame, columns=columns, show="headings", selectmode='browse', style="Treeview", height=18)
for col in columns:
    tree.heading(col, text=col, anchor='w')
    tree.column("Item", anchor='w', width=250, stretch=False)
    tree.column("LastCount", anchor='w', width=175, stretch=False)
tree.pack(expand=True, fill="both", padx=2, pady=2)  # Minimal padding

# Controls frame
controls_frame = ctk.CTkFrame(root)
controls_frame.pack(pady=2, fill="x")  # Minimal vertical padding

# Sub-frame for input and buttons
entry_controls_frame = ctk.CTkFrame(controls_frame)
entry_controls_frame.pack(pady=2, anchor="center")  # Tighter vertical packing

# "-" button
button_subtract = ctk.CTkButton(
    entry_controls_frame,
    text="-",
    command=lambda: update_count('subtract'),
    width=45,
    font=("Helvetica", 14),
    corner_radius=3
)
button_subtract.pack(side="left", padx=2)  # Minimal horizontal padding

# Entry field
entry_value = tk.Entry(
    entry_controls_frame,
    font=("Helvetica", 14),
    justify="center",
    width=5,
    validate="key",
    validatecommand=vcmd
)
entry_value.pack(side="left", padx=2)  # Minimal horizontal padding

# "+" button
button_add = ctk.CTkButton(
    entry_controls_frame,
    text="+",
    command=lambda: update_count('add'),
    width=45,
    font=("Helvetica", 14),
    corner_radius=3
)
button_add.pack(side="left", padx=2)  # Minimal horizontal padding

# Log view frame
log_view_frame = ctk.CTkFrame(root)
log_view_frame.pack(side=tk.BOTTOM, fill='both', expand=True, padx=2, pady=2)  # Minimal margins

# Log Treeview
log_view_columns = ("Timestamp", "Item", "Action", "SAN Number")
log_view = ttk.Treeview(log_view_frame, columns=log_view_columns, show="headings", style="Treeview", height=15)
for col in log_view_columns:
    log_view.heading(col, text=col, anchor='w')
    log_view.column("Timestamp", anchor='w', width=175, stretch=False)
    log_view.column("Item", anchor='w', width=180, stretch=False)
    log_view.column("Action", anchor='w', width=100, stretch=False)

scrollbar_log = ttk.Scrollbar(log_view_frame, orient="vertical", command=log_view.yview)
scrollbar_log.pack(side='right', fill='y')
log_view.configure(yscrollcommand=scrollbar_log.set)
log_view.pack(expand=True, fill='both')


# Add the copying functionality here
def add_copy_option(tree):
    def copy_selection():
        selected_item = tree.item(tree.focus())['values']
        item_text = ", ".join(map(str, selected_item))
        root.clipboard_clear()  # Clear the clipboard
        root.clipboard_append(item_text)  # Append new value to the clipboard

    # Create a menu
    context_menu = tk.Menu(tree, tearoff=0)
    context_menu.add_command(label="Copy", command=copy_selection)

    # Function to show the menu
    def show_context_menu(event):
        try:
            context_menu.tk_popup(event.x_root, event.y_root)
        finally:
            context_menu.grab_release()

    tree.bind("<Button-3>", show_context_menu)  # Bind right-click event

# Apply the add_copy_option function to the Treeviews
add_copy_option(tree)
add_copy_option(log_view)

root.after(100, update_treeview)
update_log_view()

root.mainloop()